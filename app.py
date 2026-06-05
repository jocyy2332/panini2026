"""
Panini FIFA World Cup 2026 - Web App
Ejecutar: python app.py
"""
from flask import Flask, jsonify, request, render_template, send_file
import json, os, io
try:
    import pandas as pd
    PANDAS_OK = True
except:
    PANDAS_OK = False
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_OK = True
except:
    EXCEL_OK = False

app = Flask(__name__)
SAVE_FILE = "progreso.json"

GROUPS = [
    ("A", [("MEX","México"), ("RSA","Sudáfrica"), ("KOR","Corea del Sur"), ("CZE","Chequia")]),
    ("B", [("CAN","Canadá"), ("BIH","Bosnia y Herz."), ("QAT","Catar"), ("SUI","Suiza")]),
    ("C", [("BRA","Brasil"), ("MAR","Marruecos"), ("HAI","Haití"), ("SCO","Escocia")]),
    ("D", [("USA","EEUU"), ("PAR","Paraguay"), ("AUS","Australia"), ("TUR","Turquía")]),
    ("E", [("GER","Alemania"), ("CUW","Curazao"), ("CIV","C. de Marfil"), ("ECU","Ecuador")]),
    ("F", [("NED","Países Bajos"), ("JPN","Japón"), ("SWE","Suecia"), ("TUN","Túnez")]),
    ("G", [("BEL","Bélgica"), ("EGY","Egipto"), ("IRN","Irán"), ("NZL","Nueva Zelanda")]),
    ("H", [("ESP","España"), ("CPV","Cabo Verde"), ("KSA","Arabia Saudita"), ("URU","Uruguay")]),
    ("I", [("FRA","Francia"), ("SEN","Senegal"), ("IRQ","Irak"), ("NOR","Noruega")]),
    ("J", [("ARG","Argentina"), ("ALG","Argelia"), ("AUT","Austria"), ("JOR","Jordania")]),
    ("K", [("POR","Portugal"), ("COD","RD Congo"), ("UZB","Uzbekistán"), ("COL","Colombia")]),
    ("L", [("ENG","Inglaterra"), ("CRO","Croacia"), ("GHA","Ghana"), ("PAN","Panamá")]),
    ("FW",[("FWC","FIFA World Cup"), ("CC","Coca-Cola")]),
]

SPECIAL_COUNTS = {"FWC": 19, "CC": 14}

def build_stickers():
    s = {}
    for group, teams in GROUPS:
        for code, name in teams:
            total = SPECIAL_COUNTS.get(code, 20)
            for i in range(1, total + 1):
                key = f"{code} {i}"
                s[key] = {"group": group, "team": name, "code": code, "num": i, "count": 0}
    return s

def load_data():
    stickers = build_stickers()
    if os.path.exists(SAVE_FILE):
        try:
            with open(SAVE_FILE, "r") as f:
                saved = json.load(f)
            for k, v in saved.items():
                if k in stickers:
                    stickers[k]["count"] = v
        except:
            pass
    return stickers

def save_data(stickers):
    with open(SAVE_FILE, "w") as f:
        json.dump({k: v["count"] for k, v in stickers.items()}, f)

def get_stats(stickers):
    total = len(stickers)
    have  = sum(1 for v in stickers.values() if v["count"] >= 1)
    return {
        "total":    total,
        "have":     have,
        "missing":  total - have,
        "repeated": sum(1 for v in stickers.values() if v["count"] > 1),
        "extras":   sum(max(0, v["count"]-1) for v in stickers.values()),
        "pct":      round(have / total * 100, 1) if total else 0,
    }

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/stickers")
def api_stickers():
    return jsonify(load_data())

@app.route("/api/stats")
def api_stats():
    return jsonify(get_stats(load_data()))

@app.route("/api/add", methods=["POST"])
def api_add():
    key = request.json.get("key","").strip().upper()
    stickers = load_data()
    if key not in stickers:
        return jsonify({"ok": False, "msg": f"Código '{key}' no encontrado."})
    stickers[key]["count"] += 1
    save_data(stickers)
    v = stickers[key]
    return jsonify({"ok": True, "key": key, "team": v["team"],
                    "count": v["count"], "repeated": v["count"] > 1})

@app.route("/api/remove", methods=["POST"])
def api_remove():
    key = request.json.get("key","").strip().upper()
    stickers = load_data()
    if key not in stickers or stickers[key]["count"] == 0:
        return jsonify({"ok": False, "msg": "No se puede quitar."})
    stickers[key]["count"] -= 1
    save_data(stickers)
    return jsonify({"ok": True, "key": key, "count": stickers[key]["count"]})

@app.route("/api/bulk", methods=["POST"])
def api_bulk():
    text = request.json.get("text","")
    stickers = load_data()
    added, not_found = [], []
    for line in text.replace(",", "\n").split("\n"):
        key = line.strip().upper()
        if not key: continue
        if key in stickers:
            stickers[key]["count"] += 1
            added.append(key)
        else:
            not_found.append(key)
    if added:
        save_data(stickers)
    return jsonify({"ok": True, "added": added, "not_found": not_found})

@app.route("/api/export")
def api_export():
    """Descarga el progreso como archivo JSON backup"""
    stickers = load_data()
    data = {k: v["count"] for k, v in stickers.items() if v["count"] > 0}
    buf = io.BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8"))
    buf.seek(0)
    return send_file(buf, mimetype="application/json",
                     as_attachment=True, download_name="panini2026_backup.json")

@app.route("/api/import", methods=["POST"])
def api_import():
    """Restaura el progreso desde un archivo JSON backup"""
    try:
        data = request.json.get("data", {})
        stickers = build_stickers()
        for k, v in data.items():
            if k in stickers:
                stickers[k]["count"] = int(v)
        save_data(stickers)
        stats = get_stats(stickers)
        return jsonify({"ok": True, "stats": stats})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})



@app.route("/api/import/excel", methods=["POST"])
def api_import_excel():
    if not PANDAS_OK:
        return jsonify({"ok": False, "msg": "pandas no instalado"})
    try:
        file = request.files.get("file")
        if not file:
            return jsonify({"ok": False, "msg": "No se recibio archivo"})

        import io as _io
        raw = file.read()
        stickers = build_stickers()

        try:
            df_miss = pd.read_excel(_io.BytesIO(raw), sheet_name="Faltantes")
            df_miss = df_miss[df_miss["Código"].notna()]
            missing = set(df_miss["Código"].astype(str).str.strip().tolist())
        except:
            missing = set()

        try:
            df_rep = pd.read_excel(_io.BytesIO(raw), sheet_name="Repetidas (Duplicados)")
            df_rep = df_rep[df_rep["Código"].notna() & df_rep["Código"].astype(str).str.match(r"^[A-Z]+\s+\d+$")]
            df_rep["Copias Extra"] = df_rep["Copias Extra"].fillna(1).astype(int)
        except:
            df_rep = None

        for key in stickers:
            stickers[key]["count"] = 0 if key in missing else 1

        if df_rep is not None:
            for _, row in df_rep.iterrows():
                key = str(row["Código"]).strip()
                extras = int(row["Copias Extra"])
                if key in stickers:
                    stickers[key]["count"] = 1 + extras

        save_data(stickers)
        stats = get_stats(stickers)
        return jsonify({"ok": True, "stats": stats})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

@app.route("/api/export/excel")
def api_export_excel():
    """Descarga el progreso como Excel"""
    if not EXCEL_OK:
        return jsonify({"error": "openpyxl no instalado"}), 500

    stickers = load_data()
    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ──
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22

    stats = get_stats(stickers)
    ws.append(["Panini FIFA World Cup 2026"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Tengo",     stats["have"]])
    ws.append(["Me faltan", stats["missing"]])
    ws.append(["Repetidas", stats["repeated"]])
    ws.append(["Total",     stats["total"]])
    ws.append(["% completado", f"{stats['pct']}%"])

    # ── Hoja 2: Colección completa ──
    ws2 = wb.create_sheet("Colección")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 10

    headers = ["Código", "Equipo", "Grupo", "Estado", "Copias"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A1A2E")
        cell.alignment = Alignment(horizontal="center")

    green  = PatternFill("solid", fgColor="E8F7ED")
    red    = PatternFill("solid", fgColor="FDE8E8")
    orange = PatternFill("solid", fgColor="FFF3E0")

    for key, v in stickers.items():
        count = v["count"]
        estado = "✅ Tengo" if count == 1 else ("🔁 Repetida" if count > 1 else "❌ Falta")
        row = [key, v["team"], f"Grupo {v['group']}", estado, count]
        ws2.append(row)
        fill = green if count == 1 else (orange if count > 1 else red)
        for cell in ws2[ws2.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    # ── Hoja 3: Repetidas ──
    ws3 = wb.create_sheet("Repetidas")
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 14
    ws3.append(["Código", "Equipo", "Grupo", "Copias totales", "Extras p/cambiar"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C05A00")
        cell.alignment = Alignment(horizontal="center")
    for key, v in stickers.items():
        if v["count"] > 1:
            ws3.append([key, v["team"], f"Grupo {v['group']}", v["count"], v["count"]-1])

    # ── Hoja 4: Faltantes ──
    ws4 = wb.create_sheet("Me faltan")
    ws4.column_dimensions["A"].width = 12
    ws4.column_dimensions["B"].width = 22
    ws4.column_dimensions["C"].width = 10
    ws4.append(["Código", "Equipo", "Grupo"])
    for cell in ws4[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="B83232")
        cell.alignment = Alignment(horizontal="center")
    for key, v in stickers.items():
        if v["count"] == 0:
            ws4.append([key, v["team"], f"Grupo {v['group']}"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="panini2026_reporte.xlsx")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
